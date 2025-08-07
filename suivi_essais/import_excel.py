import pandas as pd
from openpyxl import load_workbook
from .models import Essai, Affaire, Etat, Client
from django.db import IntegrityError
from datetime import datetime


def convertir_champ(valeur, type_attendu):
    try:
        if pd.isnull(valeur):
            return None
        if type_attendu == "date":
            if isinstance(valeur, datetime):
                return valeur.date()
            return pd.to_datetime(valeur, dayfirst=True).date()
        if type_attendu == "float":
            return float(valeur)
        if type_attendu == "bool":
            return str(valeur).strip().lower() in ("true", "1", "oui", "x")
        return str(valeur).strip()
    except Exception:
        return None


def importer_essais_depuis_excel(fichier_excel):
    print("Ouverture du fichier Excel...")
    wb = load_workbook(fichier_excel, data_only=True)

    if "IMPORT" not in wb.defined_names:
        raise ValueError("Le range nommé 'IMPORT' est introuvable dans ce fichier Excel.")

    sheet_name, cell_range = list(wb.defined_names["IMPORT"].destinations)[0]
    ws = wb[sheet_name]
    data = ws[cell_range]

    colonnes = [cell.value.strip().lower() if cell.value else None for cell in data[0]]
    lignes = [[cell.value for cell in row] for row in data[1:]]
    df = pd.DataFrame(lignes, columns=colonnes)

    print("Colonnes détectées :", colonnes)
    print("Nombre de lignes détectées :", len(df))
    print("Aperçu du fichier Excel :")
    print(df.head(5).to_string(index=False))

    type_mapping = {"C": "CAROTTE", "S": "SONDAGE", "P": "PERMEA"}

    champs_model = {field.name for field in Essai._meta.get_fields()}
    types_champs = {
        "identifiant": "str", "type": "str", "code_chantier": "str", "dlr": "date",
        "dprog": "date", "drea": "date", "denlev": "date", "drendu": "date",
        "commentaire": "str", "libelle": "str", "ville": "str", "rue": "str",
        "da_etiquette": "date", "da_formulaire": "date", "da_creation": "date",
        "da_clot": "date", "campagne_code_chantier": "str", "os_libelle": "str",
        "os_no_engagement": "str", "os_no": "str", "coord_lon": "float",
        "coord_lat": "float", "cc48_x": "float", "cc48_y": "float", "cc48_z": "float",
        "dleve": "date", "commentaire_saisie": "str", "etat": "str", "prevoir_enlev": "bool",
    }

    nb_ajoutes = 0
    client_ems, _ = Client.objects.get_or_create(nom="EMS")

    for i, ligne in df.iterrows():
        print(f"\nLigne {i+1} - préparation des champs...")

        identifiant = convertir_champ(ligne.get("identifiant"), "str")
        if not identifiant:
            print(" -> Ignorée : identifiant manquant")
            continue
        if Essai.objects.filter(identifiant=identifiant).exists():
            print(f" -> Ignorée : doublon d'identifiant {identifiant}")
            continue

        essai_data = {"identifiant": identifiant}

        # Conversion des champs standards
        for cle, type_val in types_champs.items():
            if cle in ("etat", "code_chantier", "type"):
                continue
            if cle in champs_model:
                essai_data[cle] = convertir_champ(ligne.get(cle), type_val)

        # Type d’essai
        type_val = convertir_champ(ligne.get("type"), "str")
        type_essai = type_mapping.get(type_val.upper(), type_val.upper())
        if type_essai not in ("CAROTTE", "SONDAGE", "PERMEA"):
            print(f" -> Ignorée : type inconnu {type_val}")
            continue
        essai_data["type_essai"] = type_essai

        # État
        etat_code = convertir_champ(ligne.get("etat"), "str")
        try:
            etat_obj = Etat.objects.get(code=int(etat_code))
            essai_data["etat"] = etat_obj
        except (Etat.DoesNotExist, ValueError):
            print(f"⚠️ Etat inconnu (code={etat_code}) — ligne ignorée")
            continue

        # Affaire
        code_affaire = convertir_champ(ligne.get("code_chantier"), "str")
        if not code_affaire:
            print("❌ Code chantier manquant — ligne ignorée")
            continue

        affaire = Affaire.objects.filter(code_affaire=code_affaire).first()
        if not affaire:
            print(f"ℹ️ Affaire {code_affaire} introuvable, tentative de création...")
            etat_defaut = Etat.objects.filter(code=1).first()
            if not etat_defaut:
                print("❌ État par défaut manquant (code = 1) — affaire non créée")
                continue
            try:
                affaire = Affaire.objects.create(
                    code_affaire=code_affaire,
                    nom_chantier=convertir_champ(ligne.get("libelle"), "str") or "",
                    adresse=" ".join([
                        convertir_champ(ligne.get("ville"), "str") or "",
                        convertir_champ(ligne.get("rue"), "str") or ""
                    ]).strip(),
                    numero_os=convertir_champ(ligne.get("os_libelle"), "str") or "",
                    date_limite=convertir_champ(ligne.get("dlr"), "date"),
                    client=client_ems,
                    localisation="",
                    commentaires="",
                    etat=etat_defaut
                )
                print(f"✅ Affaire créée : {affaire.code_affaire}")
            except Exception as e:
                print(f"❌ Erreur lors de la création de l'affaire : {e}")
                continue

        essai_data["affaire"] = affaire

        # Création de l’essai
        try:
            Essai.objects.create(**essai_data)
            print(f"✅ Essai ajouté : {identifiant}")
            nb_ajoutes += 1
        except IntegrityError as e:
            print(f"❌ Erreur d’intégrité : {e}")
        except Exception as e:
            print(f"❌ Erreur : {e}")

    print(f"\nNombre total d'essais ajoutés : {nb_ajoutes}")
    return nb_ajoutes
